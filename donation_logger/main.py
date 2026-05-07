# Activate virtual enviornment before running:
#   python3 -m venv .venv
#   source .venv/bin/activate   (Mac/Linux)
#   .venv/Scripts/activate      (Windows)
#
# Install Dependancies:
#
# Run:
#   python3 main.py
import sounddevice as sd
import wavio as wv
import numpy as np
import json
import os
from datetime import datetime
from dotenv import load_dotenv
from openai import OpenAI
from pydantic import BaseModel, Field
from enum import Enum
from datetime import date
from typing import Optional
from openpyxl import load_workbook

load_dotenv()
client = OpenAI(api_key = os.getenv("OPENAI_API_KEY"))

#Global variables
FREQ = 44100
MAX_DURATION = 60

class FoodType(Enum):
    NP = "NP"
    FF = "FF"
    O = "O"

class Entry(BaseModel):
    donation_date: date = Field(default_factory = date.today)
    food_type: FoodType = Field(description = "NP = Non perishable, FF = Fresh Food, O = Other")
    donor_name: Optional[str] = None
    weight: Optional[float] = None


def record_input():
    """
    Records audio until user hits enter. 
    Returns the str filename of the saved .wav file.
    """

    chunks = []

    def callback(indata, frames, time, status):
        #Called automatically every time a new audio chunk is ready
        chunks.append(indata.copy())

    filename = f"recording_{datetime.now().strftime('%Y%m%d_%H%M%S')}.wav"

    print("Recording... press Enter to stop.")
    with sd.InputStream(samplerate=FREQ, channels=1, callback=callback):
        input()

    recording = np.concatenate(chunks, axis=0)
    wv.write(filename, recording, FREQ, sampwidth=2)

    return filename

def convert_audio(filename):
    """Takes audio file and passes it into OpenAI's Whisper to convert it to text."""
    with open(filename, "rb") as audio_file:
        response = client.audio.transcriptions.create(model="gpt-4o-transcribe", file=audio_file)

    return response.text

def create_data(transcript):
    """Takes transcript and runs it through ChatGPT outputting a JSON object."""
    response = client.responses.parse(
        model="gpt-5.4-mini",
        input=[
            {"role": "system", "content": """You are a data extraction assistant for a food donation nonprofit.
                Extract the following fields from the donation description:
                - food_type: NP (Non-Perishable), FF (Fresh Food), or O (Other)
                - donor_name: name of the donor if mentioned
                - weight: weight in pounds if mentioned
                If a field is not mentioned, leave it as null.
                Ignore corrections, just use the final stated value (e.g. 'perishable, wait no non-perishable' = NP)."""},
            {
                "role": "user",
                "content": transcript
            }
        ],
        text_format=Entry
    )

    return response.output_parsed

def input_data(data, filename):
    """Uses the inputted data structure to add a row to the spreadsheet"""
    wb = load_workbook(filename)
    ws = wb.active

    last_row = max(
        (row for row in range(3, ws.max_row + 1) if ws.cell(row, 1).value is not None),
        default = 2
    )

    print(f"max_row: {ws.max_row}")
    print(f"last_row found: {last_row}")

    ws.cell(last_row + 1, 1).value = data.donation_date
    ws.cell(last_row + 1, 2).value = data.food_type.value
    ws.cell(last_row + 1, 6).value = data.donor_name
    ws.cell(last_row + 1, 7).value = data.weight

    wb.save(filename)

    return None

if __name__ == "__main__":
    print("=== TABLE Food Donation Logger ===")
    print("Press Enter to log a donation, or type 'q' to quit.")
    
    while True:
        user_input = input("\n> ")
        if user_input.lower() == 'q':
            print("Exiting. Goodbye!")
            break
        
        input_audio = record_input()
        transcript = convert_audio(input_audio)
        
        print(f"\nTranscript: \"{transcript}\"")
        print("Processing...")
        
        data = create_data(transcript)
        
        print("\n--- Entry Logged ---")
        print(f"  Date:   {data.donation_date}")
        print(f"  Type:   {data.food_type.value}")
        print(f"  Donor:  {data.donor_name or 'Anonymous'}")
        print(f"  Weight: {data.weight or 'Not specified'} lbs")
        
        input_data(data, "2025_Incoming_Outgoing_Food Log.xlsx")
        print("\n✓ Saved to spreadsheet!")
        print("\nPress Enter to log another donation, or type 'q' to quit.")